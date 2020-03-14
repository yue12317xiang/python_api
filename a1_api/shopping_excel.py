from openpyxl import load_workbook
from read_config import Read_Config
from pd import PD
class Doo_Excel:
    def get_data(self,file_name):
        wb = load_workbook(file_name)
        shop=eval(Read_Config().read_config("ceae.config",'SHOP','shop'))
        shopping_tel = (getattr(PD,"shop"))
        test_data=[]
        for key in shop:
            sheet = wb[key]
            if shop[key] == 'all':
                for i in range(2,sheet.max_row+1):
                    set_data = {}
                    set_data["code_id"] = sheet.cell(i,1).value
                    set_data["url"] = sheet.cell(i,2).value
                    if sheet.cell(i,3).value.find('${shopping_tel}') != -1:
                        set_data["data"] = sheet.cell(i,3).value.replace('${shopping_tel}',str(shopping_tel))
                    elif sheet.cell(i,3).value.find('${shopping_tellp}') != -1:
                         set_data["data"] = sheet.cell(i,3).value.replace('${shopping_tellp}',str(shopping_tel+100))
                    else:
                        set_data["data"] = sheet.cell(i,3).value
                    set_data["method"] = sheet.cell(i,4).value
                    set_data["title"] = sheet.cell(i,5).value
                    set_data["code"] = sheet.cell(i,6).value
                    set_data["sheet_name"] = key
                    test_data.append(set_data)
                    self.update_tel(shopping_tel+1,file_name,'init')
                    
            else:
                # print('------------1122111------------',mode[key])
                for code_id in shop[key]:
                    set_data = {}
                    set_data["code_id"] = sheet.cell(code_id+1,1).value
                    set_data["url"] = sheet.cell(code_id+1,2).value
                    if sheet.cell(code_id,3).value.find('${shopping_tel}') != -1:
                        set_data["data"] = sheet.cell(code_id,3).value.replace('${shopping_tel}',str(shopping_tel))
                    elif sheet.cell(code_id,3).value.find('${shopping_tellp}') != -1:
                         set_data["data"] = sheet.cell(code_id,3).value.replace('${shopping_tellp}',str(shopping_tel+100))
                    else:
                        set_data["data"] = sheet.cell(code_id+1,3).value
                    set_data["method"] = sheet.cell(code_id+1,4).value
                    set_data["title"] = sheet.cell(code_id+1,5).value
                    set_data["code"] = sheet.cell(code_id+1,6).value
                    set_data["sheet_name"] = key
                    test_data.append(set_data)
                    self.update_tel(shopping_tel+1,file_name,'init')
        return test_data
            
    def write_data(self,file_name,sheet_name,i,test_request,qiwang):
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        sheet.cell(i,7).value = test_request
        sheet.cell(i,8).value = qiwang
        wb.save(file_name)
    def update_tel(self,shopping_tel,filename,sheetname):
        wb = load_workbook(filename)
        sheet = wb[sheetname]
        sheet.cell(2,1).value = shopping_tel
        wb.save(filename)
        

if __name__ == '__main__':
    res = Doo_Excel().get_data("data.xlsx")
    print(res)
            